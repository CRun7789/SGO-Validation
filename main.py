import importlib
import os
import time
import pandas as pd
from src.acquisition.irs_data_download import download_and_concatenate_irs_files
from src.processing.irs_data_filter import (
    filter_by_irs_criteria,
    filter_by_name,
    filter_out_avoid_words,
    filter_by_recent_ruling_date,
)
from src.processing.sgo_scorer import compute_sgo_scores

_990n_processor = importlib.import_module("src.utils.990n_processor")


def main():
    """Main function to download, filter, score, and export IRS SGO data."""
    start_time = time.perf_counter()
    print("="*60)
    print("IRS Data Processing Pipeline")
    print("="*60)

    # Download and concatenate all IRS files
    print("\nDownloading IRS data from all regions...\n")
    irs_data = download_and_concatenate_irs_files()

    if irs_data is None:
        print("Failed to download IRS data. Exiting.")
        _print_elapsed(start_time)
        return 1

    # Apply IRS classification, name, avoid-word, and recent ruling filters
    print("\nApplying IRS classification filters...")
    irs_data = filter_by_recent_ruling_date(irs_data) # filter by date first, since that's the biggest
    print(f"\tAfter date: {len(irs_data)}")
    irs_data = filter_by_name(irs_data)
    print(f"\tAfter name (SGO words): {len(irs_data)}")
    irs_data = filter_out_avoid_words(irs_data)
    print(f"\tAfter name (words to avoid): {len(irs_data)}")
    irs_data = filter_by_irs_criteria(irs_data)
    print(f"\tAfter IRS criteria: {len(irs_data)}")
    print(f"\nAfter all filtering: {len(irs_data)} rows")

    # Apply sortingSGOs scoring logic to the already-filtered set
    print("\nComputing confidence scores...")
    scorer_results = compute_sgo_scores(irs_data)
    irs_data['sgo_scorer_score'] = scorer_results['sgo_scorer_score']
    irs_data['scoring_path']     = scorer_results['scoring_path']
    print(f"Scored {len(irs_data)} organizations")

    path_counts = irs_data['scoring_path'].value_counts()
    for path, count in path_counts.items():
        print(f"  {path}: {count}")

    # Compute weighted combined score
    #   NTEE path or CERTIFIED → 50% irs_filter_score + 50% sgo_scorer_score
    #   NO_NTEE / UNCLASSIFIED / DISQUALIFIED → 65% irs_filter_score + 35% sgo_scorer_score
    ntee_mask = irs_data['scoring_path'].isin(['NTEE', 'CERTIFIED'])

    irs_data['combined_score'] = 0.0
    irs_data.loc[ntee_mask, 'combined_score'] = (
        0.50 * irs_data.loc[ntee_mask, 'irs_filter_score'] +
        0.50 * irs_data.loc[ntee_mask, 'sgo_scorer_score']
    ).round(1)
    irs_data.loc[~ntee_mask, 'combined_score'] = (
        0.65 * irs_data.loc[~ntee_mask, 'irs_filter_score'] +
        0.35 * irs_data.loc[~ntee_mask, 'sgo_scorer_score']
    ).round(1)
    irs_data['combined_score'] -= 1 # Subtract 1 because nothing is 100% certain

    # Re-sort by combined score descending, then ruling date descending
    print("\nSorting by newest and most relevant organizations...")
    irs_data = irs_data.sort_values(
        ['combined_score', 'RULING'],
        ascending=[False, False],
        na_position='last'
    ).reset_index(drop=True)

    # Display summary
    print("\nDataset Info:")
    print(f"  Shape: {irs_data.shape}")
    print(f"\nScore Summary:")
    print(f"  IRS Filter Score  — min: {irs_data['irs_filter_score'].min()}, "
          f"max: {irs_data['irs_filter_score'].max()}, "
          f"avg: {irs_data['irs_filter_score'].mean():.1f}")
    print(f"  SGO Scorer Score  — min: {irs_data['sgo_scorer_score'].min()}, "
          f"max: {irs_data['sgo_scorer_score'].max()}, "
          f"avg: {irs_data['sgo_scorer_score'].mean():.1f}")
    print(f"  Combined Score    — min: {irs_data['combined_score'].min()}, "
          f"max: {irs_data['combined_score'].max()}, "
          f"avg: {irs_data['combined_score'].mean():.1f}")


    # Add contact info / website lookup (prefer EIN, fallback to NAME)
    print("\n\nFinding contact info through 990N form...")
    def _lookup_website(row):
        ein = row['EIN'] if pd.notna(row.get('EIN')) else None
        name = row['NAME'] if pd.notna(row.get('NAME')) else None
        site = _990n_processor.get_website_with_status(ein=ein, name=name)
        return pd.Series({'website': site})

    website_df = irs_data.apply(_lookup_website, axis=1)
    irs_data['website'] = website_df['website']

    # TODO: Add additional contact info retrieval here before exporting

    # Save outputs
    csv_dir  = os.path.join("data", "processed", "csv")
    xlsx_dir = os.path.join("data", "processed", "xlsx")
    os.makedirs(csv_dir,  exist_ok=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    # Plain CSV
    csv_file = os.path.join(csv_dir, "combined_irs_data.csv")
    irs_data.to_csv(csv_file, index=False)
    print(f"\nCSV saved to {csv_file}")

    # Excel with score columns highlighted
    xlsx_file = os.path.join(xlsx_dir, "combined_irs_data_scored.xlsx")
    excel_df = _remove_certified_rows(irs_data)
    _export_excel(excel_df, xlsx_file)
    print(f"Excel saved to {xlsx_file}")

    print("\n" + "="*60)
    print("Processing complete!")
    print("="*60)
    _print_elapsed(start_time)
    return 0


def _remove_certified_rows(df, scoring_col='scoring_path'):
    """Return a copy of `df` with rows where `scoring_col` == 'CERTIFIED' removed.
    We already know these are SGOs, so we can exclude them.

    This is applied only to the Excel output; the CSV export retains all rows.
    """
    if scoring_col in df.columns:
        return df[df[scoring_col] != 'CERTIFIED'].copy()
    return df.copy()


def _print_elapsed(start_time):
    """Print elapsed wall-clock time since `start_time` (seconds)."""
    try:
        elapsed = time.perf_counter() - start_time
        if elapsed < 60:
            print(f"Total runtime: {elapsed:.2f} seconds")
        else:
            mins, secs = divmod(int(elapsed), 60)
            hours, mins = divmod(mins, 60)
            if hours:
                print(f"Total runtime: {hours}h {mins}m {secs}s")
            else:
                print(f"Total runtime: {mins}m {secs}s")
    except Exception:
        # If something goes wrong printing elapsed time, don't raise.
        pass


def _export_excel(df, path):
    """
    Export SGO results to a formatted Excel file.

    Layout:
      Columns A-? (hidden)  : raw IRS data columns
      Columns ?-? (visible) : Region | State | Organization Name | Org Classification |
                              Business Address | Contact Name | Email | Website | Phone Number |
                              Ruling Date | Scoring Method | SGO Likelihood
      Last two (hidden)     : IRS Filter Score | SGO Scorer Score
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.filters import FilterColumn, Filters
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    # ── 1. Build business address ────────────────────────────────────────────
    def _address(row):
        parts = [str(row.get(c, '') or '').strip() for c in ('STREET', 'CITY', 'STATE', 'ZIP')]
        return ', '.join(p for p in parts if p)

    # ── 2. Define column groups ──────────────────────────────────────────────
    # Raw IRS columns to hide at the left — keep them for reference / CSV export
    raw_hidden = [
        'EIN', 'SORT_NAME', 'STREET', 'CITY', 'ZIP',
        'GROUP', 'SUBSECTION', 'AFFILIATION', 'CLASSIFICATION',
        'DEDUCTIBILITY', 'FOUNDATION', 'ACTIVITY', 'ORGANIZATION',
        'STATUS', 'TAX_PERIOD', 'ASSET_CD', 'INCOME_CD',
        'FILING_REQ_CD', 'PF_FILING_REQ_CD', 'ACCT_PD',
        'ASSET_AMT', 'INCOME_AMT', 'REVENUE_AMT',
    ]
    raw_hidden = [c for c in raw_hidden if c in df.columns]

    # Visible display columns: (output_header, source_or_value)
    # source_or_value is a column name in df, or None for blank placeholder columns
    DISPLAY = [
        ('Region',                  'region'),
        ('State',                   'STATE'),
        ('Organization Name',       'NAME'),
        ('Org Classification',      'NTEE_CD'),
        ('Business Address',        '__address__'),
        ('Contact Name',            'ICO'),
        ('Email',                   None),
        ('Website',                 'website'),
        ('Phone Number',            None),
        ('Ruling Date',             'RULING'),
        ('Scoring Method',          'scoring_path'),
        ('SGO Likelihood',          'combined_score'),
    ]

    # Score sub-columns hidden after SGO Likelihood
    SCORE_HIDDEN = [
        ('IRS Filter Score',  'irs_filter_score'),
        ('SGO Scorer Score',  'sgo_scorer_score'),
    ]

    # ── 3. Assemble output DataFrame ─────────────────────────────────────────
    out = pd.DataFrame(index=df.index)
    for col in raw_hidden:
        out[col] = df[col]

    addr_series = df.apply(_address, axis=1)
    for header, src in DISPLAY:
        if src == '__address__':
            out[header] = addr_series
        elif src is None:
            out[header] = ''
        elif src in df.columns:
            out[header] = df[src]
        else:
            out[header] = ''

    for header, src in SCORE_HIDDEN:
        out[header] = df[src] if src in df.columns else ''

    # ── 4. Write to Excel ────────────────────────────────────────────────────
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        out.to_excel(writer, index=False, sheet_name='SGO Results')
        ws = writer.sheets['SGO Results']

        all_headers   = list(out.columns)
        n_raw_hidden  = len(raw_hidden)
        display_names = {h for h, _ in DISPLAY}
        score_sub     = {h for h, _ in SCORE_HIDDEN}
        score_main    = {'SGO Likelihood'}

        # ── Styles ──────────────────────────────────────────────────────────
        hdr_fill   = PatternFill('solid', fgColor='1F4E79')   # dark blue
        hdr_font   = Font(bold=True, color='FFFFFF', size=11)
        score_fill = PatternFill('solid', fgColor='BDD7EE')   # light blue
        score_font = Font(bold=True, color='1F4E79', size=11)
        center     = Alignment(horizontal='center', vertical='center', wrap_text=False)

        # Column widths for each visible header
        COL_WIDTHS = {
            'Region':               25,
            'State':                 8,
            'Organization Name':    70,
            'Org Classification':   18,
            'Business Address':     48,
            'Contact Name':         24,
            'Email':                28,
            'Website':              28,
            'Phone Number':         16,
            'Ruling Date':          13,
            'Scoring Method':       22,
            'SGO Likelihood':       14,
            'IRS Filter Score':     15,
            'SGO Scorer Score':     15,
        }

        # ── Format each column ───────────────────────────────────────────────
        for col_idx, col_name in enumerate(all_headers, start=1):
            letter = get_column_letter(col_idx)
            dim    = ws.column_dimensions[letter]
            cell   = ws.cell(row=1, column=col_idx)

            # Header style
            if col_name in score_main | score_sub:
                cell.fill  = score_fill
                cell.font  = score_font
            else:
                cell.fill  = hdr_fill
                cell.font  = hdr_font
            cell.alignment = center

            # Width
            dim.width = COL_WIDTHS.get(col_name, 12)

            # Hide raw IRS columns, Region, and score sub-columns
            if col_idx <= n_raw_hidden or col_name == 'Region' or col_name in score_sub:
                dim.hidden = True
            
        # Highlight orgs with a score >80 as 'yes, reach out'
        rule = CellIsRule(operator='greaterThan', formula=['79.9'], fill=PatternFill(start_color='A8EBA7', end_color='A8EBA7', fill_type='solid'))
        ws.conditional_formatting.add('AI2:AI1000000', rule)

        # ── Header row height ────────────────────────────────────────────────
        ws.row_dimensions[1].height = 28

        # ── Freeze header ────────────────────────────────────────────────────
        ws.freeze_panes = 'A2'

        # ── Auto-filter across all columns so Region/State/Ruling sortable ──
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(all_headers))}1"
        )


if __name__ == "__main__":
    exit(main())
