import csv
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import time

sys.stdout.reconfigure(encoding="utf-8")
from dataclasses import asdict

from models import SGO
from sources import urls, blocked_urls, manual_sources

_HERE = Path(__file__).parent  # state_lists/ — used to resolve manual file paths
_ROOT = _HERE.parent           # repo root — needed to import from src/

# ── certified-SGO lookup (from the hand-maintained list in sgo_scorer.py) ────
# Insert repo root so "src.processing.sgo_scorer" is importable regardless of
# where the script is launched from.
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
from utils import normalize_name

# Certified-SGO lookup — built lazily on first call to filter_certified().
# Deferring the sgo_scorer import avoids pulling pandas into the startup path
# (~2 s on a cold interpreter).
_CERTIFIED_NORMALIZED: dict[str, set[str]] | None = None

def _get_certified_normalized() -> dict[str, set[str]]:
    global _CERTIFIED_NORMALIZED
    if _CERTIFIED_NORMALIZED is None:
        from src.processing.sgo_scorer import CERTIFIED_BY_STATE
        _CERTIFIED_NORMALIZED = {
            state: {normalize_name(n) for n in names}
            for state, names in CERTIFIED_BY_STATE.items()
        }
    return _CERTIFIED_NORMALIZED


from parsers.html_parser import parse_html_table, parse_html_list
from parsers.pdf_parser import parse_pdf
from parsers.file_parser import parse_xlsx, parse_csv, parse_docx
from parsers.state_parsers import parse_html_al, parse_pdf_az, parse_pdf_ks, parse_pdf_nv, parse_html_fl, parse_html_la, parse_html_sc, parse_docx_va

_STATE_PDF_PARSERS = {
    "az": parse_pdf_az,
    "ks": parse_pdf_ks,
    "nv": parse_pdf_nv,
}

# State-specific parsers for HTML page-list sources (keyed by 2-letter state code).
_STATE_HTML_PARSERS = {
    "al": parse_html_al,
    "fl": parse_html_fl,
    "la": parse_html_la,
    "sc": parse_html_sc,
}

# State-specific parsers for Word (.docx) sources (keyed by 2-letter state code).
_STATE_DOCX_PARSERS = {
    "va": parse_docx_va,
}


def get_parser(name: str):
    """
    Return the appropriate parser for a urls-dict entry.
    Key names encode format ("pdf list", "xlsx list", etc.) and state (first 2 chars).
    State-specific parsers override the generic ones for AZ, KS, NV (PDF) and FL (HTML).
    """
    n = name.lower()
    state_code = name[:2].lower()
    if "pdf" in n:
        state_parser = _STATE_PDF_PARSERS.get(state_code)
        if state_parser:
            return lambda content, state, url: state_parser(content, state, url)
        return lambda content, state, url: parse_pdf(content, state, url)
    if "xlsx" in n:
        return lambda content, state, url: parse_xlsx(content, state, url)
    if "csv" in n or "excel" in n:
        return lambda content, state, url: parse_csv(content, state, url)
    if "word" in n:
        state_parser = _STATE_DOCX_PARSERS.get(state_code)
        if state_parser:
            return lambda content, state, url: state_parser(content, state, url)
        return lambda content, state, url: parse_docx(content, state, url)
    # HTML sources — check for state-specific parser first
    html_parser = _STATE_HTML_PARSERS.get(state_code)
    if html_parser:
        return lambda content, state, url: html_parser(content.decode("utf-8", errors="replace"), state, url)
    if "page list" in n:
        return lambda content, state, url: parse_html_table(content.decode("utf-8", errors="replace"), state, url)
    return lambda content, state, url: parse_html_list(content.decode("utf-8", errors="replace"), state, url)


# ── post-processing ──────────────────────────────────────────────────────────

_CID_MAP = {
    "(cid:415)": "ti",  # 'ti' ligature in certain PDF fonts
}

_TEXT_FIXES = {
    "â€™": "'",  # UTF-8 right single quotation mark mis-decoded as Windows-1252
}

_NON_SGO_STARTS = (
    "pursuant to", "the following scholarship", "these qualified",
    "received cash donations", "scholarship organizations which",
    "scholarship organizations under",
    "organizations under", "organizations which",
    "authorized under this chapter",
    "for the 20",                               # "for the 2025 calendar year."
    "note:", "registered scholarship grant",
    "organizations (sgos)", "sto name", "sgo name",
    "school tuition organizations certified",
    "scholarship granting organizations certified",
    "scholarship organization name",            # column header "Scholarship Organization Name"
    "name address", "address city", "mailing address",
    "add me to your", "get on ",
    "a list of the educational",              # MO xlsx title row
    "contact:",                                 # "Contact: Name, Title" lines
    "approved scholarship",                     # NH PDF section header fragment
    "program year",                             # "2025-2026 Program Year" header
)

# Job-title pattern: "Firstname Lastname, Title [more words]"
_CONTACT_TITLE_RE = re.compile(
    r"^[\w\-']+ [\w\-']+,\s+[\w\s]*(CEO|President|Director|Manager|Coordinator|Founder|Administrator|Principal|Officer|Chair|Treasurer|Secretary|Staff)\b",
    re.IGNORECASE,
)

# Trailing phone number on an otherwise-valid name line, e.g. "Org Name (888)707-2465"
# Also catches vanity numbers like "(866)622-4ASK"
_TRAILING_PHONE_RE = re.compile(
    r"\s*\(?\d{3}\)?[\s.-]?[\d]{3}[\s.-][\d\w]{4}(\s*\(fax\))?$",
    re.IGNORECASE,
)

_TRAILING_URL_RE    = re.compile(r"\s+(?:www\.|https?://)\S+.*$", re.IGNORECASE)
_TRAILING_PO_RE     = re.compile(r"\s+P\.?O\.?\s+Box.*$", re.IGNORECASE)
_TRAILING_BRACKET_RE = re.compile(r"\s*\[.+\]$")
_TRAILING_SO_RE     = re.compile(r"\s+-\s+SO$", re.IGNORECASE)  # e.g. "Org Name - SO"
# Commas in org names (e.g. "Acme Scholarships, Inc.", "Acme Fund, LLC") cause
# csv.writer to wrap the field in quotes, making the CSV look inconsistent.
# Replace ", Suffix" with " Suffix" and strip any remaining bare trailing comma.
_COMMA_SUFFIX_RE    = re.compile(r",\s+(?=[A-Z])", re.IGNORECASE)  # ", Inc" → " Inc"

# ── contact-field normalisation ───────────────────────────────────────────────
_PHONE_DIGITS_RE = re.compile(r"\D")  # strip everything that isn't a digit


def _normalize_phone(raw: str | None) -> str | None:
    """
    Reformat a phone number string to (NXX) NXX-XXXX.

    Strips all non-digit characters and uses the last 10 digits (ignoring
    country-code prefixes such as "1-").  Returns None if the result does not
    have exactly 10 digits.
    """
    if not raw:
        return None
    digits = _PHONE_DIGITS_RE.sub("", raw)
    if len(digits) > 10:
        digits = digits[-10:]   # drop leading country code
    if len(digits) != 10:
        return raw              # unrecognised format — leave untouched
    return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"


def _normalize_email(raw: str | None) -> str | None:
    """Return email address in all-lowercase, or None if blank."""
    if not raw:
        return None
    return raw.strip().lower()

_NON_SGO_PATTERNS = [
    re.compile(r, re.IGNORECASE) for r in [
        r"^\(?\d{4}[-–]\d{4}\)?\s*(\w.*)?$",                 # year range: 2025-2026 or "2025-2026 Program Year"
        r"^\(updated .+\)$",                                # (updated April, 2026)
        r"^\(?\d{3}\)?[\s.-]?\d{3}[\s.-]\d{4}$",           # bare phone number
        r"^[\w.+\-]+@[\w\-]+\.[a-z]{2,}$",                 # bare email address
        r"^https?://",                                      # bare URL
        r"^www\.",                                          # bare URL
        # Street address: starts with a number + optional direction + street type
        r"^\d{1,6}[,.\s]+(W\.?|E\.?|N\.?|S\.?|West|East|North|South)?\s*.{0,60}\b(Street|St\b|Ave\b|Avenue|Blvd|Boulevard|Rd\b|Road|Dr\b|Drive|Way\b|Lane\b|Ln\b|Pkwy|Parkway|Circle|Ct\b|Court|Suite|Ste\b)\b",
        # Line with embedded email (contact-person lines like "Name, Title email@org")
        r"\S+@\S+\.\w{2,}",
        # Line ending with "(fax)" — always a contact line
        r"\(fax\)$",
        # "City, ST 12345" pattern (city/state/zip line)
        r"\b[A-Z]{2}\s+\d{5}\b",
        # Website embedded in what looks like contact info: ends with ".org" or ".com" alone
        r"\b\w+\.(org|com|net|edu)\s*$",
        # Single structural word (e.g. "ORGANIZATION", "Organization") — never an org name alone
        r"^(Organization|Scholarship|Foundation|Institute|Fund|Association|Corporation|Society)s?$",
        r"^Scholarship Organizations?$",           # RI PDF column header
    ]
]


def _fix_cids(text: str) -> str:
    for cid, replacement in _CID_MAP.items():
        text = text.replace(cid, replacement)
    for bad, good in _TEXT_FIXES.items():
        text = text.replace(bad, good)
    return text


def _is_sgo(name: str) -> bool:
    """Return True if name looks like an actual org, not a header/footer/contact line."""
    t = name.strip()
    if len(t) < 4:
        return False
    tl = t.lower()
    if any(tl.startswith(p) for p in _NON_SGO_STARTS):
        return False
    if _CONTACT_TITLE_RE.match(t):
        return False
    if any(pat.search(t) for pat in _NON_SGO_PATTERNS):
        return False
    return True


def postprocess(sgos: list[SGO]) -> list[SGO]:
    """
    Clean the extracted SGO list before writing:
      1. Replace known PDF encoding artifacts.
      2. Remove entries that are obviously not org names.
      3. Deduplicate within each state.
    """
    cleaned: list[SGO] = []
    seen: set[tuple[str, str]] = set()
    removed = 0

    for sgo in sgos:
        name = _fix_cids(sgo.name)
        if not _is_sgo(name):
            removed += 1
            continue
        if name.isupper() and " " in name:
            # Title-case word by word.  Words of ≤3 characters keep their
            # original casing (preserving abbreviations and acronyms such as
            # "LLC", "SGO", "of", "1"), EXCEPT common articles/conjunctions
            # ("the", "and") which are always title-cased.  Longer words get
            # first-letter-upper + rest-lower, which correctly handles
            # apostrophes ("CHILDREN’S" → "Children’s") without the
            # str.title() bug ("Children’S").
            _ALWAYS_TITLE = {"the", "and"}
            name = " ".join(
                word[0].upper() + word[1:].lower()
                if len(word) > 3 or word.lower() in _ALWAYS_TITLE
                else word
                for word in name.split(" ")
            )
        name = _TRAILING_BRACKET_RE.sub("", name).strip()
        name = _TRAILING_SO_RE.sub("", name).strip()
        name = _TRAILING_PHONE_RE.sub("", name).strip()
        name = _TRAILING_URL_RE.sub("", name).strip()
        name = _TRAILING_PO_RE.sub("", name).strip()
        name = _COMMA_SUFFIX_RE.sub(" ", name).rstrip(",").strip()
        if not _is_sgo(name):
            removed += 1
            continue
        # Normalise the dedup key: strip a leading "the " so "The Org" and
        # "Org" (without the article) aren't treated as different organisations.
        name_key = name.lower()
        if name_key.startswith("the "):
            name_key = name_key[4:]
        key = (sgo.state, name_key)
        if key in seen:
            continue
        seen.add(key)
        sgo.name  = name
        sgo.phone = _normalize_phone(sgo.phone)
        sgo.email = _normalize_email(sgo.email)
        cleaned.append(sgo)

    print(f"\nPost-processing: removed {removed} non-SGO entries, "
          f"{len(sgos) - removed - len(cleaned)} duplicates → {len(cleaned)} records remain")
    return cleaned


def mark_certified(sgos: list[SGO]) -> list[SGO]:
    """
    Set sgo.known = "Yes" for every org already present in CERTIFIED_BY_STATE.

    All records are kept; the 'Known?' column in the output lets the reviewer
    see at a glance which orgs are already validated SGOs.

    Matching uses normalize_name() so differences in articles, corporate
    suffixes, punctuation, and capitalisation are ignored.
    """
    marked = 0
    for sgo in sgos:
        certified = _get_certified_normalized().get(sgo.state.lower(), set())
        if normalize_name(sgo.name) in certified:
            sgo.known = "Yes"
            marked += 1
    print(f"Certified check: marked {marked} already-certified orgs as Known")
    return sgos


# ── output ───────────────────────────────────────────────────────────────────

_FIELDNAMES = ["state", "name", "ein", "address", "phone", "email", "website", "raw_source", "known"]


def write_results(sgos: list[SGO], path: str) -> None:
    """Write SGO records to a CSV and a matching Excel file."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_FIELDNAMES)
        writer.writeheader()
        writer.writerows(asdict(s) for s in sgos)
    print(f"Wrote {len(sgos)} SGO records to {path}")

    xlsx_path = str(_XLSX_DIR / Path(path).with_suffix(".xlsx").name)
    _XLSX_DIR.mkdir(parents=True, exist_ok=True)
    _export_xlsx(sgos, xlsx_path)


def _export_xlsx(sgos: list[SGO], path: str) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_FIELDNAMES)

    for sgo in sgos:
        ws.append([getattr(sgo, f) or "" for f in _FIELDNAMES])

    # Remove any trailing empty rows (openpyxl can leave one after the last append)
    while ws.max_row > 1:
        if all(cell.value in (None, "") for cell in ws[ws.max_row]):
            ws.delete_rows(ws.max_row)
        else:
            break

    # Column widths in character units, matching _FIELDNAMES order:
    # state, name, ein, address, phone, email, website, raw_source, known
    _COL_WIDTHS = [6, 58, 10, 53, 13, 38, 33, 71, 8]
    for i, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    while True:
        try:
            wb.save(path)
            print(f"Wrote {len(sgos)} SGO records to {path}")
            break
        except PermissionError:
            print(f"\n[ERROR] Cannot write {Path(path).name} — the file is open in another program.")
            try:
                input("Close the file in Excel, then press Enter to retry... ")
            except EOFError:
                print("Skipping xlsx export (non-interactive mode).")
                break


# ── pipeline entry points ────────────────────────────────────────────────────

def check_urls() -> None:
    """
    Check every URL in the registry for reachability and report its status,
    cross-referencing blocked_urls and manual_sources to catch:

      - Newly blocked URLs not yet in blocked_urls
      - Previously blocked URLs that have come back online
      - Manual files that are no longer needed because the URL is now accessible
      - Blocked URLs that have a manual file covering them (expected, informational)
      - Blocked URLs with no fallback at all (need attention)
    """
    from fetcher import check_site  # deferred — requests is slow to import
    print(f"\nChecking {len(urls)} URLs ...\n")
    for name, url in urls.items():
        print(f"-----------\nChecking {name}")
        result = check_site(url)
        status  = f"[{result.status}]" if result.status else "[---]"
        is_blocked  = bool(result.blocked)   # True or None both mean not fully OK
        is_ok       = result.blocked is False
        in_blocklist = name in blocked_urls
        has_manual   = name in manual_sources
        manual_path  = (_HERE / manual_sources[name]) if has_manual else None
        manual_exists = has_manual and manual_path.exists()

        if is_ok and not in_blocklist and not has_manual:
            # Normal case — nothing to flag
            print(f"✅ OK     {status}  {url}")

        elif is_ok and not in_blocklist and has_manual:
            # URL is live but a manual file exists — the file is now redundant
            print(f"✅ OK     {status}  {url}")
            print(f"  ⚠️  Manual file exists but URL is accessible — remove {manual_path}")

        elif is_ok and in_blocklist and not has_manual:
            # URL has recovered — blocked_urls entry is stale
            print(f"✅ OK     {status}  {url}")
            print(f"  🔄 URL is now accessible — remove '{name}' from blocked_urls in sources.py")

        elif is_ok and in_blocklist and has_manual:
            # URL has recovered AND a manual file exists — both entries are now redundant
            print(f"✅ OK     {status}  {url}")
            print(f"  🔄 URL is now accessible — remove '{name}' from blocked_urls in sources.py")
            print(f"  ⚠️  Manual file is also redundant — remove {manual_path}")

        elif not is_ok and not in_blocklist and not has_manual:
            # Newly blocked URL — not in blocklist, no fallback
            print(f"🚨 NEWLY BLOCKED  {status}  {url}  —  {result.reason}")
            print(f"  Add '{name}' to blocked_urls in sources.py")

        elif not is_ok and not in_blocklist and has_manual:
            # Newly blocked but manual file covers it — still flag so blocklist gets updated
            print(f"🚨 NEWLY BLOCKED  {status}  {url}  —  {result.reason}")
            print(f"  Add '{name}' to blocked_urls in sources.py")
            if manual_exists:
                print(f"  Manual file is covering it: {manual_path}")

        elif not is_ok and in_blocklist and not has_manual:
            # Expected block, no manual fallback — informational
            print(f"🚫 BLOCKED  {status}  {url}  —  {result.reason}  (known)")

        elif not is_ok and in_blocklist and has_manual:
            # Expected block, manual file in use
            if manual_exists:
                print(f"🚫 BLOCKED  {status}  {url}  —  {result.reason}  (known, manual file in use: {manual_path.name})")
            else:
                print(f"🚫 BLOCKED  {status}  {url}  —  {result.reason}  (known)")
                print(f"  ⚠️  Manual file is missing — download and place at {manual_path}")
                print(f"  See manual/README.md for instructions.")

        time.sleep(1)


_MAX_FETCH_WORKERS = 8  # concurrent HTTP threads; all sources are different domains


_OUTPUT_DIR = _ROOT / "data" / "processed"
_CSV_DIR    = _OUTPUT_DIR / "csv"
_XLSX_DIR   = _OUTPUT_DIR / "xlsx"

def run(output_path: str = str(_CSV_DIR / "state_sgo_lists.csv"),
        full_pipeline: bool = False) -> None:
    """
    Fetch and parse every accessible data source; write results to output_path.

    Skips:
    - "page source" keys (navigation/program pages, not data sources)
    - blocked_urls entries that have no corresponding manual file

    For blocked_urls entries that DO have a manual file in manual_sources,
    reads the local file instead of attempting a network request.

    Network fetches run concurrently (up to _MAX_FETCH_WORKERS threads) so the
    total wall-clock time is dominated by the slowest single source rather than
    the sum of all round-trips.  Parsing runs sequentially after all fetches
    complete so parse output is ordered and readable.
    """
    import subprocess
    from fetcher import fetch_bytes  # deferred — requests is slow to import

    _start = time.perf_counter()
    _last_checkpoint = [_start]  # list so the nested closure can mutate it

    def _fmt(secs: float) -> str:
        if secs < 60:
            return f"{secs:.1f}s"
        m, s = divmod(int(secs), 60)
        return f"{m}m {s}s"

    def _checkpoint(label: str, since: float = None) -> None:
        """Print section time (since last checkpoint or explicit `since`) and running total."""
        now = time.perf_counter()
        section = now - (since if since is not None else _last_checkpoint[0])
        total   = now - _start
        print(f"  ⏱  {label}: {_fmt(section)} elapsed  (total {_fmt(total)})")
        _last_checkpoint[0] = now

    # ── 1. Build work list ────────────────────────────────────────────────────
    # Each item: (name, url, manual_path_or_None)
    work: list[tuple[str, str, Path | None]] = []

    for name, url in urls.items():
        if "page source" in name.lower():
            continue
        if name in blocked_urls:
            if name not in manual_sources:
                print(f"[SKIP] {name} — blocked ({blocked_urls[name]})")
                continue
            manual_path = _HERE / manual_sources[name]
            if not manual_path.exists():
                print(f"[SKIP] {name} — blocked; manual file not found ({manual_path})")
                print( "         See manual/README.md for download instructions.")
                continue
            work.append((name, url, manual_path))
        else:
            work.append((name, url, None))

    # ── 2. Fetch all sources concurrently ─────────────────────────────────────
    def _fetch(item: tuple[str, str, Path | None]) -> tuple[str, str, bytes | None, str | None]:
        name, url, manual_path = item
        if manual_path is not None:
            return name, url, manual_path.read_bytes(), None
        try:
            return name, url, fetch_bytes(url), None
        except Exception as e:
            return name, url, None, str(e)

    n_network = sum(1 for _, _, mp in work if mp is None)
    n_manual  = len(work) - n_network
    print(f"\nFetching {n_network} remote source(s) in parallel"
          f"{f' + {n_manual} manual file(s)' if n_manual else ''} ...")

    fetched: list[tuple[str, str, bytes | None, str | None]] = []
    with ThreadPoolExecutor(max_workers=_MAX_FETCH_WORKERS) as pool:
        future_to_item = {pool.submit(_fetch, item): item for item in work}
        for future in as_completed(future_to_item):
            name, url, content, err = future.result()
            label = "[MANUAL]" if future_to_item[future][2] is not None else f"[{name[:2].upper()}]"
            if err:
                print(f"  {label} {name} — FAILED: {err}")
            else:
                print(f"  {label} {name} — {len(content):,} bytes")
            fetched.append((name, url, content, err))
    _checkpoint("Fetch")

    # ── 3. Parse sequentially (fast; keeps output readable and ordered) ────────
    fetched.sort(key=lambda r: r[0])  # deterministic order regardless of arrival

    all_sgos: list[SGO] = []
    for name, url, content, err in fetched:
        if err or content is None:
            continue
        state = name[:2].upper()
        parser = get_parser(name)
        print(f"\n[{state}] Parsing {name} ...")
        try:
            sgos = parser(content, state, url)
            print(f"  → {len(sgos)} SGOs extracted")
            all_sgos.extend(sgos)
        except Exception as e:
            print(f"  FAILED: {e}")
    _checkpoint("Parse")

    all_sgos = postprocess(all_sgos)
    all_sgos = mark_certified(all_sgos)
    _checkpoint("Post-process + certified filter")

    _CSV_DIR.mkdir(parents=True, exist_ok=True)
    write_results(all_sgos, output_path)
    _checkpoint("Write state_sgo_lists")

    print("\nState lists pipeline complete.")

    if not full_pipeline:
        xlsx_path = _XLSX_DIR / Path(output_path).with_suffix(".xlsx").name
        try:
            answer = input("\nOpen state_sgo_lists.xlsx in Excel? [y/N]: ").strip().lower()
        except EOFError:
            answer = ""
        if answer == "y":
            subprocess.Popen(
                ["explorer.exe", str(xlsx_path)],
                creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
            )
            print("Opening file...")
            time.sleep(2)
        return

    # ── Combination pipeline ──────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Starting combination pipeline (IRS EO BMF)...")
    print("=" * 60)
    result = subprocess.run(
        [sys.executable, str(_ROOT / "main.py")],
        cwd=str(_ROOT),
    )
    if result.returncode != 0:
        print(f"\nCombination pipeline exited with code {result.returncode}.")
    _checkpoint("Combination pipeline (IRS EO BMF)")

    # ── Enrich state_sgo_lists with IRS data ──────────────────────────────────
    _ENRICHED_FIELDNAMES = [
        "state", "name", "ein", "phone", "email", "website",
        "Contact Name", "address", "Ruling Date", "raw_source", "known",
    ]
    state_sgo_path = _CSV_DIR  / "state_sgo_lists.csv"
    combined_path  = _CSV_DIR  / "combined_irs_data.csv"
    enriched_path  = _CSV_DIR  / "enriched_data.csv"

    if not combined_path.exists():
        print(f"\n[WARN] {combined_path.name} not found — skipping enrichment.")
    else:
        # Build two indexes from combined_irs_data:
        #   irs_name_index — (STATE_upper, normalized_name) → best IRS row
        #   irs_ein_index  — EIN (digits only, no dashes) → best IRS row
        # When duplicates exist, keep the row with the highest combined_score.
        def _best_row(existing: dict, candidate: dict) -> dict:
            try:
                return candidate if (
                    float(candidate.get("combined_score") or 0)
                    > float(existing.get("combined_score") or 0)
                ) else existing
            except (ValueError, TypeError):
                return existing

        irs_name_index: dict[tuple[str, str], dict] = {}
        irs_ein_index:  dict[str, dict] = {}
        with open(combined_path, encoding="utf-8") as f:
            for irs_row in csv.DictReader(f):
                state = irs_row.get("STATE", "").strip().upper()
                name_key = (state, normalize_name(irs_row.get("NAME", "")))
                irs_name_index[name_key] = _best_row(
                    irs_name_index[name_key], irs_row
                ) if name_key in irs_name_index else irs_row

                ein_raw = re.sub(r"\D", "", irs_row.get("EIN", ""))
                if ein_raw:
                    irs_ein_index[ein_raw] = _best_row(
                        irs_ein_index[ein_raw], irs_row
                    ) if ein_raw in irs_ein_index else irs_row

        def _irs_address(irs_row: dict) -> str:
            parts = [irs_row.get(c, "").strip()
                     for c in ("STREET", "CITY", "STATE", "ZIP")]
            return ", ".join(p for p in parts if p)

        # ── Pass 1: IRS enrichment (fast — in-memory lookups) ────────────────
        total = matched = 0
        enriched_rows: list[dict] = []
        with open(state_sgo_path, encoding="utf-8") as fin:
            for row in csv.DictReader(fin):
                total += 1
                enriched = {**row, "Contact Name": "", "Ruling Date": ""}
                # Primary lookup: (state, normalized name)
                name_key = (row.get("state", "").strip().upper(),
                            normalize_name(row.get("name", "")))
                irs = irs_name_index.get(name_key)
                # Secondary lookup: EIN (if the state list already has one)
                if irs is None:
                    ein_raw = re.sub(r"\D", "", row.get("ein", "") or "")
                    if ein_raw:
                        irs = irs_ein_index.get(ein_raw)
                if irs:
                    matched += 1
                    if not enriched.get("ein"):
                        enriched["ein"]          = irs.get("EIN", "").strip()
                    if not enriched.get("address"):
                        enriched["address"]      = _irs_address(irs)
                    if not enriched.get("phone"):
                        enriched["phone"]        = irs.get("phone", "").strip()
                    if not enriched.get("email"):
                        enriched["email"]        = irs.get("email", "").strip()
                    if not enriched.get("website"):
                        enriched["website"]      = irs.get("website", "").strip()
                    enriched["Contact Name"]     = irs.get("ICO", "").strip()
                    enriched["Ruling Date"]      = irs.get("RULING", "").strip()
                enriched_rows.append(enriched)

        print(f"\nEnrichment: matched {matched}/{total} rows from IRS data.")
        _checkpoint("IRS enrichment")

        # ── Write enriched CSV ────────────────────────────────────────────────
        with open(enriched_path, "w", newline="", encoding="utf-8") as fout:
            writer = csv.DictWriter(fout, fieldnames=_ENRICHED_FIELDNAMES)
            writer.writeheader()
            writer.writerows(enriched_rows)
        print(f"Wrote enriched data → {enriched_path.name}")

        # ── Export enriched_data.xlsx ─────────────────────────────────────────
        import openpyxl
        enriched_xlsx = _XLSX_DIR / "enriched_data.xlsx"
        _XLSX_DIR.mkdir(parents=True, exist_ok=True)
        _ENRICHED_COL_HEADERS = {
            "state":        "State",
            "name":         "Name",
            "ein":          "EIN",
            "phone":        "Phone",
            "email":        "Email",
            "website":      "Website",
            "Contact Name": "Contact Name",
            "address":      "Address",
            "Ruling Date":  "Ruling Date",
            "raw_source":   "Raw Source",
            "known":        "Known?",
        }
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([_ENRICHED_COL_HEADERS.get(f, f) for f in _ENRICHED_FIELDNAMES])
        with open(enriched_path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                ws.append([row.get(col, "") or "" for col in _ENRICHED_FIELDNAMES])
        # Remove any trailing empty rows
        while ws.max_row > 1:
            if all(cell.value in (None, "") for cell in ws[ws.max_row]):
                ws.delete_rows(ws.max_row)
            else:
                break
        # Column widths matching _ENRICHED_FIELDNAMES order:
        # State, Name, EIN, Phone, Email, Website, Contact Name, Address, Ruling Date, Raw Source, Known?
        _ENRICHED_COL_WIDTHS = [7, 44, 13, 15, 32, 32, 20, 54, 12, 67, 8]
        for i, width in enumerate(_ENRICHED_COL_WIDTHS, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        while True:
            try:
                wb.save(enriched_xlsx)
                print(f"Wrote enriched_data.xlsx to {enriched_xlsx}")
                _checkpoint("Export enriched_data.xlsx")
                break
            except PermissionError:
                print(f"\n[ERROR] Cannot write enriched_data.xlsx — the file is open in another program.")
                try:
                    input("Close the file in Excel, then press Enter to retry... ")
                except EOFError:
                    print("Skipping xlsx export (non-interactive mode).")
                    break

        # ── Offer to open in Excel ────────────────────────────────────────────
        try:
            answer = input("\nOpen enriched_data.xlsx in Excel? [y/N]: ").strip().lower()
        except EOFError:
            answer = ""
        if answer == "y":
            subprocess.Popen(
                ["explorer.exe", str(enriched_xlsx)],
                creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
            )
            print("Opening file...")
            time.sleep(2)

if __name__ == "__main__":
    run(full_pipeline=True)
